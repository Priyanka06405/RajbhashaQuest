from flask import Flask, render_template, request, redirect, session, send_file
import random
import sqlite3
import os
import io
import qrcode

from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4


app = Flask(__name__)

app.secret_key = "RajbhashaQuest2026"


# -----------------------------
# Database
# -----------------------------

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

DATABASE = os.path.join(BASE_DIR, "database.db")


def get_connection():

    conn = sqlite3.connect(DATABASE)

    conn.row_factory = sqlite3.Row

    return conn


def create_tables():

    conn = get_connection()

    cur = conn.cursor()

    # Players
    cur.execute("""
    CREATE TABLE IF NOT EXISTS players(

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        employee_id TEXT UNIQUE,

        name TEXT NOT NULL,

        department TEXT NOT NULL,

        score INTEGER DEFAULT 0,

        coins INTEGER DEFAULT 0,

        xp INTEGER DEFAULT 0,

        level INTEGER DEFAULT 1

    )
    """)

    # Questions
    cur.execute("""
    CREATE TABLE IF NOT EXISTS questions(

        id INTEGER PRIMARY KEY AUTOINCREMENT,

        question TEXT,

        option1 TEXT,

        option2 TEXT,

        option3 TEXT,

        option4 TEXT,

        answer TEXT,

        category TEXT

    )
    """)

    conn.commit()


    # Add sample questions only if empty

    count = cur.execute(
        "SELECT COUNT(*) FROM questions"
    ).fetchone()[0]


    if count == 0:

        sample = [

            (
                "Computer का हिन्दी शब्द क्या है?",
                "संगणक",
                "पत्र",
                "बैठक",
                "कर्मचारी",
                "संगणक",
                "Rajbhasha"
            ),

            (
                "File का हिन्दी शब्द क्या है?",
                "संचिका",
                "पत्र",
                "सूची",
                "दस्तावेज",
                "संचिका",
                "Rajbhasha"
            ),

            (
                "Meeting का हिन्दी शब्द क्या है?",
                "बैठक",
                "कार्यालय",
                "रिपोर्ट",
                "अनुभाग",
                "बैठक",
                "Rajbhasha"
            ),

            (
                "Office का हिन्दी शब्द क्या है?",
                "कार्यालय",
                "विद्यालय",
                "संगणक",
                "दस्तावेज",
                "कार्यालय",
                "Rajbhasha"
            ),

            (
                "Department का हिन्दी शब्द क्या है?",
                "विभाग",
                "अनुभाग",
                "संचिका",
                "बैठक",
                "विभाग",
                "Rajbhasha"
            ),

            (
                "Letter का हिन्दी शब्द क्या है?",
                "पत्र",
                "रिपोर्ट",
                "दस्तावेज",
                "संचिका",
                "पत्र",
                "Rajbhasha"
            ),

            (
                "Document का हिन्दी शब्द क्या है?",
                "दस्तावेज",
                "पत्र",
                "बैठक",
                "विभाग",
                "दस्तावेज",
                "Rajbhasha"
            ),

            (
                "Employee का हिन्दी शब्द क्या है?",
                "कर्मचारी",
                "अधिकारी",
                "अनुभाग",
                "कार्यालय",
                "कर्मचारी",
                "Rajbhasha"
            ),

            (
                "Training का हिन्दी शब्द क्या है?",
                "प्रशिक्षण",
                "बैठक",
                "दस्तावेज",
                "रिपोर्ट",
                "प्रशिक्षण",
                "Rajbhasha"
            ),

            (
                "Report का हिन्दी शब्द क्या है?",
                "प्रतिवेदन",
                "पत्र",
                "सूची",
                "दस्तावेज",
                "प्रतिवेदन",
                "Rajbhasha"
            )

        ]


        cur.executemany("""
        INSERT INTO questions(
            question,
            option1,
            option2,
            option3,
            option4,
            answer,
            category
        )
        VALUES(?,?,?,?,?,?,?)
        """, sample)

        conn.commit()


    conn.close()


# -----------------------------
# Home
# -----------------------------

@app.route("/")
def home():

    return render_template("index.html")


# -----------------------------
# Login
# -----------------------------

@app.route("/login", methods=["GET", "POST"])
def login():

    if request.method == "POST":

        employee = request.form["employee"].strip()

        name = request.form["name"].strip()

        department = request.form["department"].strip()


        if not employee or not name or not department:

            return render_template(
                "login.html",
                error="Please fill all fields."
            )


        conn = get_connection()


        player = conn.execute(
            "SELECT * FROM players WHERE employee_id=?",
            (employee,)
        ).fetchone()


        if player is None:

            conn.execute("""
            INSERT INTO players(
                employee_id,
                name,
                department
            )
            VALUES(?,?,?)
            """,
            (
                employee,
                name,
                department
            ))

            conn.commit()


        conn.close()


        session["employee"] = employee

        session["name"] = name


        return redirect("/dashboard")


    return render_template("login.html")


# -----------------------------
# Dashboard
# -----------------------------

@app.route("/dashboard")
def dashboard():

    if "employee" not in session:

        return redirect("/login")


    conn = get_connection()


    player = conn.execute(
        "SELECT * FROM players WHERE employee_id=?",
        (session["employee"],)
    ).fetchone()


    conn.close()


    return render_template(
        "dashboard.html",
        player=player
    )


# -----------------------------
# Logout
# -----------------------------

@app.route("/logout")
def logout():

    session.clear()

    return redirect("/")


# ============================================================
# QUIZ
# ============================================================

@app.route("/quiz")
def quiz():

    if "employee" not in session:

        return redirect("/login")


    conn = get_connection()


    # Start a new quiz

    if "quiz_questions" not in session:

        rows = conn.execute(
            "SELECT id FROM questions"
        ).fetchall()


        ids = [
            row["id"]
            for row in rows
        ]


        random.shuffle(ids)


        # Maximum 10 questions

        session["quiz_questions"] = ids[:10]

        session["current_question"] = 0

        session["quiz_score"] = 0


    index = session["current_question"]


    # Quiz finished

    if index >= len(session["quiz_questions"]):

        final_score = session["quiz_score"]


        conn.execute("""
        UPDATE players

        SET score = score + ?

        WHERE employee_id = ?

        """,
        (
            final_score,
            session["employee"]
        ))


        conn.commit()

        conn.close()


        session.pop("quiz_questions", None)

        session.pop("current_question", None)

        session.pop("quiz_score", None)


        return render_template(
            "result.html",
            message="🎉 Quiz Completed!",
            correct=f"Your Score: {final_score}"
        )


    question_id = session["quiz_questions"][index]


    question = conn.execute(
        "SELECT * FROM questions WHERE id=?",
        (question_id,)
    ).fetchone()


    conn.close()


    return render_template(
        "quiz.html",
        question=question,
        current=index + 1,
        total=len(session["quiz_questions"])
    )


# -----------------------------
# Check Answer
# -----------------------------

@app.route("/check_answer", methods=["POST"])
def check_answer():

    if "employee" not in session:

        return redirect("/login")


    question_id = request.form.get("question_id")

    selected = request.form.get("answer", "")


    if not question_id:

        return redirect("/quiz")


    conn = get_connection()


    question = conn.execute(
        "SELECT * FROM questions WHERE id=?",
        (question_id,)
    ).fetchone()


    if question is None:

        conn.close()

        return redirect("/quiz")


    # Correct answer

    if selected and selected == question["answer"]:

        session["quiz_score"] += 10


        # XP and coins

        conn.execute("""
        UPDATE players

        SET
            xp = xp + 5,
            coins = coins + 2

        WHERE employee_id=?

        """,
        (session["employee"],))


        conn.commit()


    # Move to next question

    session["current_question"] += 1


    conn.close()


    return redirect("/quiz")


# ============================================================
# LEADERBOARD
# ============================================================

@app.route("/leaderboard")
def leaderboard():

    conn = get_connection()


    players = conn.execute("""
        SELECT *

        FROM players

        ORDER BY score DESC, xp DESC
    """).fetchall()


    conn.close()


    return render_template(
        "leaderboard.html",
        players=players
    )


# ============================================================
# ADMIN LOGIN
# ============================================================

@app.route("/admin", methods=["GET", "POST"])
def admin():

    if request.method == "POST":

        username = request.form["username"]

        password = request.form["password"]


        if username == "admin" and password == "admin123":

            session["admin"] = True

            return redirect("/admin/dashboard")


        return render_template(
            "admin_login.html",
            error="Invalid Username or Password"
        )


    return render_template("admin_login.html")


# ============================================================
# ADMIN DASHBOARD
# ============================================================

@app.route("/admin/dashboard")
def admin_dashboard():

    if "admin" not in session:

        return redirect("/admin")


    conn = get_connection()


    total_players = conn.execute(
        "SELECT COUNT(*) FROM players"
    ).fetchone()[0]


    total_questions = conn.execute(
        "SELECT COUNT(*) FROM questions"
    ).fetchone()[0]


    conn.close()


    return render_template(
        "admin_dashboard.html",
        players=total_players,
        questions=total_questions
    )


# ============================================================
# QUESTION MANAGEMENT
# ============================================================

@app.route("/questions")
def questions():

    if "admin" not in session:

        return redirect("/admin")


    conn = get_connection()


    questions_list = conn.execute("""
        SELECT *

        FROM questions

        ORDER BY id DESC
    """).fetchall()


    conn.close()


    return render_template(
        "questions.html",
        questions=questions_list
    )


# -----------------------------
# Add Question
# -----------------------------

@app.route("/add_question", methods=["GET", "POST"])
def add_question():

    if "admin" not in session:

        return redirect("/admin")


    if request.method == "POST":

        question = request.form["question"]

        option1 = request.form["option1"]

        option2 = request.form["option2"]

        option3 = request.form["option3"]

        option4 = request.form["option4"]

        answer = request.form["answer"]

        category = request.form.get(
            "category",
            "Rajbhasha"
        )


        conn = get_connection()


        conn.execute("""
        INSERT INTO questions(
            question,
            option1,
            option2,
            option3,
            option4,
            answer,
            category
        )
        VALUES(?,?,?,?,?,?,?)
        """,
        (
            question,
            option1,
            option2,
            option3,
            option4,
            answer,
            category
        ))


        conn.commit()

        conn.close()


        return redirect("/questions")


    return render_template("add_question.html")


# -----------------------------
# Edit Question
# -----------------------------

@app.route("/edit_question/<int:question_id>",
           methods=["GET", "POST"])
def edit_question(question_id):

    if "admin" not in session:

        return redirect("/admin")


    conn = get_connection()


    question = conn.execute(
        "SELECT * FROM questions WHERE id=?",
        (question_id,)
    ).fetchone()


    if question is None:

        conn.close()

        return redirect("/questions")


    if request.method == "POST":

        conn.execute("""
        UPDATE questions

        SET
            question=?,
            option1=?,
            option2=?,
            option3=?,
            option4=?,
            answer=?,
            category=?

        WHERE id=?
        """,
        (
            request.form["question"],
            request.form["option1"],
            request.form["option2"],
            request.form["option3"],
            request.form["option4"],
            request.form["answer"],
            request.form.get(
                "category",
                "Rajbhasha"
            ),
            question_id
        ))


        conn.commit()

        conn.close()


        return redirect("/questions")


    conn.close()


    return render_template(
        "edit_question.html",
        question=question
    )


# -----------------------------
# Delete Question
# -----------------------------

@app.route("/delete_question/<int:question_id>")
def delete_question(question_id):

    if "admin" not in session:

        return redirect("/admin")


    conn = get_connection()


    conn.execute(
        "DELETE FROM questions WHERE id=?",
        (question_id,)
    )


    conn.commit()

    conn.close()


    return redirect("/questions")


# ============================================================
# EXCEL QUESTION UPLOAD
# ============================================================

@app.route("/upload_questions",
           methods=["GET", "POST"])
def upload_questions():

    if "admin" not in session:

        return redirect("/admin")


    if request.method == "POST":

        file = request.files.get("file")


        if not file or file.filename == "":

            return render_template(
                "upload_questions.html",
                message="Please select an Excel file."
            )


        try:

            workbook = load_workbook(file)

            sheet = workbook.active


            conn = get_connection()

            imported = 0


            # First row = headings

            for row in sheet.iter_rows(
                min_row=2,
                values_only=True
            ):

                if not row[0]:

                    continue


                question = str(row[0])

                option1 = str(row[1])

                option2 = str(row[2])

                option3 = str(row[3])

                option4 = str(row[4])

                answer = str(row[5])

                category = (
                    str(row[6])
                    if len(row) > 6 and row[6]
                    else "Rajbhasha"
                )


                conn.execute("""
                INSERT INTO questions(
                    question,
                    option1,
                    option2,
                    option3,
                    option4,
                    answer,
                    category
                )
                VALUES(?,?,?,?,?,?,?)
                """,
                (
                    question,
                    option1,
                    option2,
                    option3,
                    option4,
                    answer,
                    category
                ))


                imported += 1


            conn.commit()

            conn.close()


            return render_template(
                "upload_questions.html",
                message=(
                    f"✅ {imported} questions "
                    "imported successfully!"
                )
            )


        except Exception as e:

            return render_template(
                "upload_questions.html",
                message=f"❌ Error: {e}"
            )


    return render_template(
        "upload_questions.html"
    )


# ============================================================
# ADMIN REPORTS
# ============================================================

@app.route("/admin/reports")
def admin_reports():

    if "admin" not in session:

        return redirect("/admin")


    conn = get_connection()


    total_players = conn.execute(
        "SELECT COUNT(*) FROM players"
    ).fetchone()[0]


    total_questions = conn.execute(
        "SELECT COUNT(*) FROM questions"
    ).fetchone()[0]


    total_score = conn.execute(
        "SELECT COALESCE(SUM(score),0) FROM players"
    ).fetchone()[0]


    total_coins = conn.execute(
        "SELECT COALESCE(SUM(coins),0) FROM players"
    ).fetchone()[0]


    top_players = conn.execute("""
        SELECT *

        FROM players

        ORDER BY score DESC, xp DESC

        LIMIT 10
    """).fetchall()


    conn.close()


    return render_template(
        "admin_reports.html",
        total_players=total_players,
        total_questions=total_questions,
        total_score=total_score,
        total_coins=total_coins,
        top_players=top_players
    )


# ============================================================
# PLAYER MANAGEMENT
# ============================================================

@app.route("/admin/players")
def admin_players():

    if "admin" not in session:

        return redirect("/admin")


    conn = get_connection()


    players = conn.execute("""
        SELECT *

        FROM players

        ORDER BY score DESC, xp DESC
    """).fetchall()


    conn.close()


    return render_template(
        "admin_players.html",
        players=players
    )


# ============================================================
# CERTIFICATE
# ============================================================

@app.route("/certificate")
def certificate():

    if "employee" not in session:

        return redirect("/login")


    conn = get_connection()


    player = conn.execute("""
        SELECT *

        FROM players

        WHERE employee_id=?
    """,
    (session["employee"],)).fetchone()


    conn.close()


    if not player:

        return redirect("/login")


    buffer = io.BytesIO()


    pdf = canvas.Canvas(
        buffer,
        pagesize=A4
    )


    width, height = A4


    pdf.setFont(
        "Helvetica-Bold",
        28
    )


    pdf.drawCentredString(
        width / 2,
        height - 150,
        "CERTIFICATE OF ACHIEVEMENT"
    )


    pdf.setFont(
        "Helvetica",
        18
    )


    pdf.drawCentredString(
        width / 2,
        height - 210,
        "RAJBHASHA QUEST"
    )


    pdf.setFont(
        "Helvetica",
        14
    )


    pdf.drawCentredString(
        width / 2,
        height - 260,
        "This certificate is proudly presented to"
    )


    pdf.setFont(
        "Helvetica-Bold",
        24
    )


    pdf.drawCentredString(
        width / 2,
        height - 320,
        player["name"]
    )


    pdf.setFont(
        "Helvetica",
        14
    )


    pdf.drawCentredString(
        width / 2,
        height - 370,
        "Employee ID: " +
        str(player["employee_id"])
    )


    pdf.drawCentredString(
        width / 2,
        height - 400,
        "Department: " +
        str(player["department"])
    )


    pdf.drawCentredString(
        width / 2,
        height - 450,
        "Score: " +
        str(player["score"])
    )


    pdf.drawCentredString(
        width / 2,
        height - 480,
        "XP: " +
        str(player["xp"])
    )


    pdf.drawCentredString(
        width / 2,
        height - 510,
        "Coins: " +
        str(player["coins"])
    )


    pdf.setFont(
        "Helvetica-Oblique",
        12
    )


    pdf.drawCentredString(
        width / 2,
        100,
        "Rajbhasha Quest Achievement Certificate"
    )


    pdf.save()


    buffer.seek(0)


    return send_file(
        buffer,
        as_attachment=True,
        download_name="Rajbhasha_Quest_Certificate.pdf",
        mimetype="application/pdf"
    )


# ============================================================
# QR CODE
# ============================================================

@app.route("/qr_code")
def qr_code():

    if "admin" not in session:

        return redirect("/admin")


    # Live website address

    url = request.host_url


    qr = qrcode.make(url)


    buffer = io.BytesIO()


    qr.save(
        buffer,
        format="PNG"
    )


    buffer.seek(0)


    return send_file(
        buffer,
        mimetype="image/png"
    )


# ============================================================
# MAIN
# ============================================================

if __name__ == "__main__":

    create_tables()

    app.run(
        debug=True,
        host="0.0.0.0"
    )
